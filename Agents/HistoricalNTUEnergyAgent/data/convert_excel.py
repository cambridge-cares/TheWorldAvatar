
'''
This python file does 3 things:
1. Converts the header of the .xlsx file
Below is the header format in the original NTU_Energy_Consumption.xlsx
'NEC_P(KW)'
'NEC_Q(KVAR)'
'CANTEEN_2_P(KW)'

The header format should be converted to below format before our agent can retrieve and map the data
NEC_P_KW
NEC_Q_KVAR
CANTEEN_2_P_KW

2. Convert the time into the format which can be read by our agent
From 1/1/2020  1:00:00 am
To   2020-01-01T01:00:00

3. Divide all values in the excel file by 10
'''

'''

'''
import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook(filename='NTU_Energy_Consumption.xlsx')
ws = wb.active

# Insert four new columns after the first column
ws.insert_cols(2, 4)

# Set the headers for the new columns
ws.cell(row=1, column=2, value="GENERATOR_NODE_P_KW")
ws.cell(row=1, column=3, value="GENERATOR_NODE_Q_KVAR")
ws.cell(row=1, column=4, value="CONNECTION_NODE_P_KW")
ws.cell(row=1, column=5, value="CONNECTION_NODE_Q_KVAR")

# Iterate over the cells in the first row and replace any parentheses with underscores
for cell in ws[1]:
    cell.value = str(cell.value).replace('(', '_').replace(')', '')

# Iterate over the rows starting from the second row
for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    # Divide all cell values in the row except the first column by 10^4
    divided_row = []
    for col_number, cell_value in enumerate(row, start=1):
        if col_number == 1:
            divided_row.append(cell_value)
        else:
            if cell_value is not None:  # check if cell value is not None
                divided_row.append(cell_value / 10000)
            else:
                divided_row.append(None)
    # Convert the date/time string to a datetime object
    datetime_obj = datetime.strptime(str(row[0]), '%Y-%m-%d %H:%M:%S')
    # Format the datetime object to a string in the desired format
    formatted_date_string = datetime_obj.strftime("%Y-%m-%dT%H:%M:%S")
    # Update the cell with the formatted date string and the divided values
    ws.cell(row=row_number, column=1, value=formatted_date_string)
    ws.cell(row=row_number, column=2, value=0)
    ws.cell(row=row_number, column=3, value=0)
    ws.cell(row=row_number, column=4, value=0)
    ws.cell(row=row_number, column=5, value=0)
    for col_number, cell_value in enumerate(divided_row[1:], start=2):
        if cell_value is not None:  # check if cell value is not None
            ws.cell(row=row_number, column=col_number, value=cell_value)
        else:
            ws.cell(row=row_number, column=col_number, value=None)

# Save the updated workbook to a file
wb.save('NTU_Energy_Consumption.xlsx')
print("Successfully converted the .xslx file!")


